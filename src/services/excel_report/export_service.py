"""
JobExportService — v2
=====================
Correcciones respecto a v1:
  - Filtro de fechas: AND dentro del rango por tipo, OR entre tipos
  - Job_type muestra solo el valor del enum ("PTL", no "JobType.PTL")
  - Subcontractor usa campo Organization
  - Colores pastel neutrales
  - Eliminados: multipliers, payment_units
  - Nuevo: CommissionDetail.Sell_Mgmt como "Selling Commission" / "Mgmt Commission"
  - Nuevo: Purchases → Purchase 1, Purchase 2... (Total_spending)
  - Nuevo: EstimateCosts → EC 1 Quantity, EC 1 Unit, etc.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select
from sqlalchemy import and_, or_, select as sa_select
from sqlalchemy.orm import selectinload

from .export_schema import JobExportColumns, JobExportFilters, JobExportRequest
from src.models.JobModel import Job
from src.models.ClientModel import Client
from src.models.MemberModel import Member
from src.models.SubcontractorModel import Subcontractor
from src.models.OrderModel import Order
from src.models.ChangeOrderModel import ChangeOrder
from src.models.ComDetailModel import CommissionDetail
from src.models.ComGroupModel import CommissionGroup
from src.models.PurchaseModel import Purchase
from src.models.EstimateCostModel import EstimateCost
from src.models.link_models.JobMember import JobMemberLink
from src.models.link_models.JobSubcontractor import JobSubcontractorLink


# ─────────────────────────────────────────────────────────────────────────────
#  Paleta pastel neutral
# ─────────────────────────────────────────────────────────────────────────────

_FONT_HEADER = Font(name="Arial", bold=True, color="3D3D3D", size=10)
_FONT_BODY = Font(name="Arial", size=10)

# Encabezados: pastel suave con texto oscuro
_FILL_JOB = PatternFill("solid", fgColor="D6E4F0")  # azul pálido
_FILL_CLIENT = PatternFill("solid", fgColor="E8F5E9")  # verde muy suave
_FILL_MEM = PatternFill("solid", fgColor="EDE7F6")  # lavanda suave
_FILL_SUB = PatternFill("solid", fgColor="FFF3E0")  # naranja muy pálido
_FILL_ORD = PatternFill("solid", fgColor="FFF8E1")  # amarillo muy pálido
_FILL_CO = PatternFill("solid", fgColor="FFFDE7")  # amarillo casi blanco
_FILL_COM = PatternFill("solid", fgColor="FCE4EC")  # rosa pálido
_FILL_PUR = PatternFill("solid", fgColor="E0F2F1")  # verde agua pálido
_FILL_EC = PatternFill("solid", fgColor="F3E5F5")  # lila muy pálido

_ALIGN_CENTER = Alignment(
    horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, list):
        return ", ".join(str(v) for v in value if v is not None)
    # Enum: devolver solo el .value
    if hasattr(value, "value"):
        return str(value.value)
    return str(value)


def _fmt_job_type(value: Any) -> str:
    """Extrae solo el nombre del tipo: 'JobType.PTL' → 'PTL'."""
    raw = str(value)
    return raw.split(".")[-1] if "." in raw else raw


def _apply_header(cell, fill: PatternFill) -> None:
    cell.font = _FONT_HEADER
    cell.fill = fill
    cell.alignment = _ALIGN_CENTER


# ─────────────────────────────────────────────────────────────────────────────
#  Consultas a BD
# ─────────────────────────────────────────────────────────────────────────────

def _query_jobs(session: Session, filters: JobExportFilters, cols: JobExportColumns) -> List[Job]:
    # Portal (sub/técnico): solo sus jobs; staff pasa intacto. Solo se llama desde
    # POST /jobs_excel/export, dentro de una petición.
    from src.utils.middleware.auth.routes_protection import scope_jobs_statement
    query = scope_jobs_statement(select(Job))

    # ── Optimizaciones (Eager Loading) ────────────────────────────────────────
    if cols.include_client:
        query = query.options(selectinload(Job.client))
    if cols.include_members:
        query = query.options(selectinload(Job.members))
    if cols.include_subcontractors:
        # Nota: aquí también cargamos la tabla intermedia si es necesario
        query = query.options(selectinload(Job.subcontractors))
    if cols.include_commissions:
        # Cargamos comdetails y TAMBIÉN su comgroup relacionado para evitar N+1
        query = query.options(
            selectinload(Job.comdetails).selectinload(CommissionDetail.comgroup)
        )
    if cols.include_purchases:
        query = query.options(selectinload(Job.purchases))
    if cols.include_estimate_costs:
        query = query.options(selectinload(Job.estimate_costs))

    # ── Filtros ───────────────────────────────────────────────────────────────
    if filters.statuses:
        query = query.where(Job.Job_status.in_(filters.statuses))

    if filters.job_types:
        query = query.where(Job.Job_type.in_(filters.job_types))

    if filters.member_ids:
        # FIX: Evitamos .distinct() sobre toda la fila para no fallar con columnas JSON.
        # Usamos IN subquery sobre el ID primario.
        sub = sa_select(JobMemberLink.job_id).where(
            JobMemberLink.member_id.in_(filters.member_ids))
        query = query.where(Job.ID_Jobs.in_(sub))

    if filters.client_id:
        query = query.where(Job.ID_Client == filters.client_id)

    if (filters.parent_mgmt_co_id):
        query = query.where(Job.client.has(
            Client.ID_Community_Tracking == filters.parent_mgmt_co_id))

    if filters.search:
        pattern = f"%{filters.search}%"
        query = query.where(or_(
            Job.Project_name.ilike(pattern),
            Job.ID_Jobs.ilike(pattern),
            Job.Project_location.ilike(pattern),
            Job.Job_status.ilike(pattern),
            Job.Service_type.ilike(pattern),
            Job.client.has(Client.Client_Community.ilike(pattern)),
            Job.client.has(Client.parent_mgmt_co.has(or_(
                ParentMgmtCo.Property_mgmt_co.ilike(pattern),
                ParentMgmtCo.Company_abbrev.ilike(pattern)
            ))),
            Job.members.any(Member.Member_Name.ilike(pattern))
        ))

    # ── Filtro de fechas corregido ────────────────────────────────────────────
    if filters.date_from or filters.date_to:
        active_types = [t.upper()
                        for t in (filters.job_types or ["QID", "PAR", "PTL"])]
        type_clauses = []

        if "PTL" in active_types:
            ptl = []
            if filters.date_from:
                ptl.append(Job.Estimated_start_date >= filters.date_from)
            if filters.date_to:
                ptl.append(Job.Estimated_start_date <= filters.date_to)
            if ptl:
                type_clauses.append(and_(*ptl))

        if any(t in active_types for t in ("QID", "PAR")):
            qp = []
            if filters.date_from:
                qp.append(Job.Date_assigned >= filters.date_from)
            if filters.date_to:
                qp.append(Job.Date_assigned <= filters.date_to)
            if qp:
                type_clauses.append(and_(*qp))

        if type_clauses:
            query = query.where(or_(*type_clauses))

    return list(session.exec(query).all())


def _get_members_for_job(session: Session, job_id: str) -> List[Tuple[str, str]]:
    """Retorna [(rol, member_name), …]."""
    rows = session.exec(
        select(JobMemberLink, Member)
        .join(Member, Member.ID_Member == JobMemberLink.member_id)
        .where(JobMemberLink.job_id == job_id)
        .order_by(JobMemberLink.rol)
    ).all()
    return [(link.rol, member.Member_Name or "") for link, member in rows]


def _get_subcontractors_for_job(session: Session, job_id: str) -> List[Subcontractor]:
    rows = session.exec(
        select(Subcontractor)
        .join(JobSubcontractorLink,
              Subcontractor.ID_Subcontractor == JobSubcontractorLink.subcontr_id)
        .where(JobSubcontractorLink.job_id == job_id)
        .order_by(Subcontractor.Organization)
    ).all()

    for r in rows:
        if r.Organization:
            r.Organization = r.Organization.replace(
                '"', '').replace('{', '').replace('}', '')

    return list(rows)


def _get_orders_for_job_sub(session: Session, job: Job, sub_id: str) -> List[Order]:
    rows = session.exec(
        select(Order).where(
            and_(
                Order.ID_Subcontractor == sub_id,
                Order.job_podio_id == job.podio_item_id,
            )
        ).order_by(Order.ID_Order)
    ).all()
    return list(rows)


def _get_change_orders_for_order(
    session: Session, job_id: str, order_id: str
) -> List[ChangeOrder]:
    rows = session.exec(
        select(ChangeOrder).where(
            and_(
                ChangeOrder.ID_Jobs == job_id,
                ChangeOrder.ID_Order == order_id,
            )
        ).order_by(ChangeOrder.ID_ChangeOrder)
    ).all()
    return list(rows)


def _get_client_name(session: Session, client_id: Optional[str]) -> str:
    if not client_id:
        return ""
    client = session.get(Client, client_id)
    return (client.Client_Community or "") if client else ""


def _get_commissions_for_job(
    session: Session, job_id: str
) -> Dict[str, Optional[float]]:
    """
    Retorna {"Selling Commission": valor, "Mgmt Commission": valor}.
    Determina el tipo por CommissionGroup.Rol:
      - Rol contiene "Selling" → Selling Commission
      - Rol contiene "Mgmt"    → Mgmt Commission
    Si hay varios registros para el mismo tipo, se suma.
    """
    result = {"Selling Commission": None, "Mgmt Commission": None}

    rows = session.exec(
        select(CommissionDetail, CommissionGroup)
        .join(CommissionGroup,
              CommissionDetail.ID_ComGroup == CommissionGroup.ID_ComGroup)
        .where(CommissionDetail.ID_Jobs == job_id)
    ).all()

    for detail, group in rows:
        rol = (group.Rol or "").lower()
        val = detail.Sell_Mgmt
        if val is None:
            continue
        if "selling" in rol:
            result["Selling Commission"] = (
                result["Selling Commission"] or 0) + val
        elif "mgmt" in rol:
            result["Mgmt Commission"] = (result["Mgmt Commission"] or 0) + val

    return result


def _get_purchases_for_job(session: Session, job_id: str) -> List[Purchase]:
    rows = session.exec(
        select(Purchase)
        .where(Purchase.ID_Jobs == job_id)
        .order_by(Purchase.ID_Purchase)
    ).all()
    return list(rows)


def _get_estimate_costs_for_job(session: Session, job_id: str) -> List[EstimateCost]:
    rows = session.exec(
        select(EstimateCost)
        .where(EstimateCost.ID_Jobs == job_id)
        .order_by(EstimateCost.ID_EstimateCost)
    ).all()
    return list(rows)


# ─────────────────────────────────────────────────────────────────────────────
#  Recolección de todos los datos
# ─────────────────────────────────────────────────────────────────────────────

def _collect_all_data(
    session: Session,
    jobs: List[Job],
    cols: JobExportColumns,
) -> Tuple[List[Dict], Dict]:
    roles_seen: Dict[str, int] = {}
    max_subs = 0
    max_orders_per_sub: List[int] = []
    max_co_per_order: List[List[int]] = []
    max_purchases = 0
    max_estimate_costs = 0

    rows = []

    # Pre-cargar miembros en lote
    job_members_map = {}
    if cols.include_members and jobs:
        job_ids = [j.ID_Jobs for j in jobs]
        member_rows = session.exec(
            select(JobMemberLink, Member)
            .join(Member, Member.ID_Member == JobMemberLink.member_id)
            .where(JobMemberLink.job_id.in_(job_ids))
            .order_by(JobMemberLink.job_id, JobMemberLink.rol)
        ).all()
        for link, member in member_rows:
            if link.job_id not in job_members_map:
                job_members_map[link.job_id] = []
            job_members_map[link.job_id].append(
                (link.rol, member.Member_Name or ""))

    # Pre-cargar órdenes en lote
    job_orders_map = {}
    if cols.include_subcontractors and jobs:
        podio_ids = [j.podio_item_id for j in jobs if j.podio_item_id]
        order_rows = session.exec(
            select(Order)
            .where(Order.job_podio_id.in_(podio_ids))
            .order_by(Order.ID_Order)
        ).all()
        for o in order_rows:
            key = (o.job_podio_id, o.ID_Subcontractor)
            if key not in job_orders_map:
                job_orders_map[key] = []
            job_orders_map[key].append(o)

    # Pre-cargar Change Orders en lote
    order_cos_map = {}
    if cols.include_subcontractors and jobs:
        job_ids = [j.ID_Jobs for j in jobs]
        co_rows = session.exec(
            select(ChangeOrder)
            .where(ChangeOrder.ID_Jobs.in_(job_ids))
            .order_by(ChangeOrder.ID_ChangeOrder)
        ).all()
        for co in co_rows:
            key = (co.ID_Jobs, co.ID_Order)
            if key not in order_cos_map:
                order_cos_map[key] = []
            order_cos_map[key].append(co)

    for job in jobs:
        row: Dict = {"_job": job}

        if cols.include_client:
            row["_client_name"] = (
                job.client.Client_Community or "") if job.client else ""

        if cols.include_members:
            members = job_members_map.get(job.ID_Jobs, [])
            row["_members"] = members
            from collections import Counter
            role_count = Counter(r for r, _ in members)
            for rol, cnt in role_count.items():
                roles_seen[rol] = max(roles_seen.get(rol, 0), cnt)

        if cols.include_subcontractors:
            subs = job.subcontractors
            sub_data = []
            for i, sub in enumerate(subs):
                if sub.Organization:
                    sub.Organization = sub.Organization.replace(
                        '"', '').replace('{', '').replace('}', '')

                key = (job.podio_item_id, sub.ID_Subcontractor)
                orders = job_orders_map.get(key, [])
                orders_data = []
                for j, order in enumerate(orders):
                    cos = order_cos_map.get((job.ID_Jobs, order.ID_Order), [])
                    orders_data.append({"order": order, "change_orders": cos})

                    if i >= len(max_co_per_order):
                        max_co_per_order.append([])
                    if j >= len(max_co_per_order[i]):
                        max_co_per_order[i].append(0)
                    max_co_per_order[i][j] = max(
                        max_co_per_order[i][j], len(cos))

                sub_data.append({"sub": sub, "orders": orders_data})

                if i >= len(max_orders_per_sub):
                    max_orders_per_sub.append(0)
                max_orders_per_sub[i] = max(max_orders_per_sub[i], len(orders))

            max_subs = max(max_subs, len(subs))
            row["_subs"] = sub_data

        if cols.include_commissions:
            # Usar relación precargada 'comdetails'
            # Necesitamos traer también el 'group' para el rol
            commissions = {"Selling Commission": None, "Mgmt Commission": None}
            for detail in job.comdetails:
                # Usar comgroup (no group)
                group = detail.comgroup
                rol = (group.Rol or "").lower() if group else ""
                val = detail.Sell_Mgmt
                if val is None:
                    continue
                if "selling" in rol:
                    commissions["Selling Commission"] = (commissions["Selling Commission"] or 0) + val
                elif "mgmt" in rol:
                    commissions["Mgmt Commission"] = (commissions["Mgmt Commission"] or 0) + val
            row["_commissions"] = commissions

        if cols.include_purchases:
            purchases = job.purchases
            row["_purchases"] = purchases
            max_purchases = max(max_purchases, len(purchases))

        if cols.include_estimate_costs:
            ecs = job.estimate_costs
            row["_estimate_costs"] = ecs
            max_estimate_costs = max(max_estimate_costs, len(ecs))

        rows.append(row)

    metadata = {
        "roles_seen":         roles_seen,
        "max_subs":           max_subs,
        "max_orders_per_sub": max_orders_per_sub,
        "max_co_per_order":   max_co_per_order,
        "max_purchases":      max_purchases,
        "max_estimate_costs": max_estimate_costs,
    }
    return rows, metadata


# ─────────────────────────────────────────────────────────────────────────────
#  Esquema de columnas dinámico
# ─────────────────────────────────────────────────────────────────────────────

def _build_column_schema(cols: JobExportColumns, metadata: Dict) -> List[Dict]:
    schema = []

    # El Excel se arma columna a columna desde el modelo: NO pasa por
    # `serialize_job` ni por `add_relationships`, asi que la redaccion central
    # no lo alcanzaba. Medido: el fichero que se descarga un subcontratista
    # llevaba los mismos valores financieros que el del Full Admin —los
    # centinelas 4444 (Gqm_formula_pricing), 9999 (Gqm_final_sold_pricing) y
    # 1414 (Acc_receivable) aparecian en los dos.
    #
    # Un export es la peor forma de filtrar el margen: sale de la aplicacion en
    # un fichero que ya no controla nadie.
    from src.utils.portal_redaction import (CAMPOS_FINANCIEROS_JOB,
                                            CAMPOS_PODIO, llamante_es_portal)
    vetados = ((CAMPOS_FINANCIEROS_JOB | CAMPOS_PODIO)
               if llamante_es_portal() else frozenset())

    # ── Campos básicos ────────────────────────────────────────────────────────
    for field in cols.basic_fields:
        f = field.value
        if f in vetados:
            continue
        if f == "Job_type":
            def key_fn(r): return _fmt_job_type(
                getattr(r["_job"], "Job_type", None))
        else:
            def key_fn(r, _f=f): return _fmt(getattr(r["_job"], _f, None))
        schema.append({"header": f, "fill": _FILL_JOB, "key": key_fn})

    # ── Cliente ───────────────────────────────────────────────────────────────
    if cols.include_client:
        schema.append({
            "header": "Client",
            "fill":   _FILL_CLIENT,
            "key": lambda r: r.get("_client_name", ""),
        })

    # ── Members (por rol) ─────────────────────────────────────────────────────
    if cols.include_members:
        for rol, max_count in sorted(metadata["roles_seen"].items()):
            for idx in range(max_count):
                header = rol if max_count == 1 else f"{rol} {idx + 1}"
                schema.append({
                    "header": header,
                    "fill":   _FILL_MEM,
                    "key":    _make_member_key(rol, idx),
                })

    # ── Subcontractors → Orders → Change Orders ───────────────────────────────
    if cols.include_subcontractors:
        max_subs = metadata["max_subs"]
        max_orders_ps = metadata["max_orders_per_sub"]
        max_co_po = metadata["max_co_per_order"]

        for s_idx in range(max_subs):
            schema.append({
                "header": f"Subcontractor {s_idx + 1}",
                "fill":   _FILL_SUB,
                "key":    _make_sub_name_key(s_idx),
            })

            n_orders = max_orders_ps[s_idx] if s_idx < len(
                max_orders_ps) else 0
            for o_idx in range(n_orders):
                for field, label in [
                    ("Title",            "Title"),
                    ("Formula",          "Formula"),
                    ("Adj_formula",      "Adj Formula"),
                    ("Notes",            "Notes"),
                    ("Ptl_hd_materials", "HD Materials"),
                ]:
                    schema.append({
                        "header": f"S{s_idx+1} Order {o_idx+1} {label}",
                        "fill":   _FILL_ORD,
                        "key":    _make_order_field_key(s_idx, o_idx, field),
                    })

                n_cos = (
                    max_co_po[s_idx][o_idx]
                    if s_idx < len(max_co_po) and o_idx < len(max_co_po[s_idx])
                    else 0
                )
                for c_idx in range(n_cos):
                    for field, label in [
                        ("Name",              "Name"),
                        ("ChangeOrderFormula", "Formula"),
                        ("State",             "State"),
                    ]:
                        schema.append({
                            "header": f"S{s_idx+1} O{o_idx+1} CO{c_idx+1} {label}",
                            "fill":   _FILL_CO,
                            "key":    _make_co_field_key(s_idx, o_idx, c_idx, field),
                        })

    # ── Comisiones ────────────────────────────────────────────────────────────
    if cols.include_commissions:
        for label in ("Selling Commission", "Mgmt Commission"):
            schema.append({
                "header": label,
                "fill":   _FILL_COM,
                "key":    _make_commission_key(label),
            })

    # ── Purchases ─────────────────────────────────────────────────────────────
    if cols.include_purchases:
        for p_idx in range(metadata["max_purchases"]):
            schema.append({
                "header": f"Purchase {p_idx + 1} Total Spending",
                "fill":   _FILL_PUR,
                "key":    _make_purchase_key(p_idx),
            })

    # ── Estimate Costs ────────────────────────────────────────────────────────
    if cols.include_estimate_costs:
        for e_idx in range(metadata["max_estimate_costs"]):
            n = e_idx + 1
            for field, label in [
                ("Quatity",      "Quantity"),
                ("Unit",         "Unit"),
                ("Unit_cost",    "Unit Cost"),
                ("Builder_cost", "Builder Cost"),
                ("Client_price", "Client Price"),
            ]:
                schema.append({
                    "header": f"EC {n} {label}",
                    "fill":   _FILL_EC,
                    "key":    _make_ec_field_key(e_idx, field),
                })

    return schema


# ── Closures de acceso ────────────────────────────────────────────────────────

def _make_member_key(rol: str, idx: int):
    def key(r):
        names = [name for r2, name in r.get("_members", []) if r2 == rol]
        return names[idx] if idx < len(names) else ""
    return key


def _make_sub_name_key(s_idx: int):
    def key(r):
        subs = r.get("_subs", [])
        if s_idx >= len(subs):
            return ""
        sub = subs[s_idx]["sub"]
        # Usa Organization; si está vacío cae en Name
        return _fmt(sub.Organization or sub.Name)
    return key


def _make_order_field_key(s_idx: int, o_idx: int, field: str):
    def key(r):
        subs = r.get("_subs", [])
        if s_idx >= len(subs):
            return ""
        orders = subs[s_idx]["orders"]
        if o_idx >= len(orders):
            return ""
        return _fmt(getattr(orders[o_idx]["order"], field, None))
    return key


def _make_co_field_key(s_idx: int, o_idx: int, c_idx: int, field: str):
    def key(r):
        subs = r.get("_subs", [])
        if s_idx >= len(subs):
            return ""
        orders = subs[s_idx]["orders"]
        if o_idx >= len(orders):
            return ""
        cos = orders[o_idx]["change_orders"]
        if c_idx >= len(cos):
            return ""
        return _fmt(getattr(cos[c_idx], field, None))
    return key


def _make_commission_key(label: str):
    def key(r):
        val = r.get("_commissions", {}).get(label)
        return _fmt(val)
    return key


def _make_purchase_key(p_idx: int):
    def key(r):
        purchases = r.get("_purchases", [])
        if p_idx >= len(purchases):
            return ""
        return _fmt(purchases[p_idx].Total_spending)
    return key


def _make_ec_field_key(e_idx: int, field: str):
    def key(r):
        ecs = r.get("_estimate_costs", [])
        if e_idx >= len(ecs):
            return ""
        return _fmt(getattr(ecs[e_idx], field, None))
    return key


# ─────────────────────────────────────────────────────────────────────────────
#  Generación del Excel
# ─────────────────────────────────────────────────────────────────────────────

def _write_excel(rows: List[Dict], column_schema: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs Export"

    # Encabezados
    for col_idx, col_def in enumerate(column_schema, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
        _apply_header(cell, col_def["fill"])
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            len(col_def["header"]) + 4, 14
        )

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Datos
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_def in enumerate(column_schema, start=1):
            value = col_def["key"](row)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _FONT_BODY
            cell.alignment = _ALIGN_LEFT
        ws.row_dimensions[row_idx].height = 18

    # Auto-ajuste de ancho (máx 45)
    for col_idx in range(1, len(column_schema) + 1):
        col_letter = get_column_letter(col_idx)
        current_w = ws.column_dimensions[col_letter].width
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value or ""
            current_w = max(current_w, min(len(str(val)), 45))
        ws.column_dimensions[col_letter].width = current_w + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
#  Punto de entrada público
# ─────────────────────────────────────────────────────────────────────────────

def generate_jobs_excel(session: Session, request: JobExportRequest) -> bytes:
    jobs = _query_jobs(session, request.filters, request.columns)

    if not jobs:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs Export"
        ws["A1"] = "No se encontraron jobs con los filtros aplicados."
        ws["A1"].font = Font(name="Arial", italic=True, color="888888")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    rows, metadata = _collect_all_data(session, jobs, request.columns)
    column_schema = _build_column_schema(request.columns, metadata)
    return _write_excel(rows, column_schema)

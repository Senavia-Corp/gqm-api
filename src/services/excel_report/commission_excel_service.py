from __future__ import annotations

import io
from datetime import datetime
from typing import Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select
from sqlalchemy.orm import selectinload

from src.models.CommissionModel import Commission
from src.models.ComGroupModel import CommissionGroup
from src.models.ComDetailModel import CommissionDetail
from src.models.JobModel import Job
from src.models.ReimbursementModel import Reimbursement

# ─── Styles ──────────────────────────────────────────────────────────────────

_FONT_HEADER = Font(name="Arial", bold=True, color="3D3D3D", size=10)
_FONT_SECTION = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_FONT_BODY = Font(name="Arial", size=10)

_FILL_SUMMARY = PatternFill("solid", fgColor="D6E4F0")   # blue
_FILL_QID_SELL = PatternFill("solid", fgColor="FCE4EC")  # pink
_FILL_QID_MGMT = PatternFill("solid", fgColor="EDE7F6")  # lavender
_FILL_PAR_MGMT = PatternFill("solid", fgColor="E8F5E9")  # green
_FILL_SEC_QID_SELL = PatternFill("solid", fgColor="E91E63")
_FILL_SEC_QID_MGMT = PatternFill("solid", fgColor="7E57C2")
_FILL_SEC_PAR_MGMT = PatternFill("solid", fgColor="43A047")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "value"):
        return str(value.value)
    return str(value)


def _hdr(ws, row: int, col: int, text: str, fill: PatternFill) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _FONT_HEADER
    cell.fill = fill
    cell.alignment = _ALIGN_CENTER


def _body(ws, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _FONT_BODY
    cell.alignment = _ALIGN_LEFT


def _section_header(ws, row: int, text: str, num_cols: int, fill: PatternFill) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _FONT_SECTION
    cell.fill = fill
    cell.alignment = _ALIGN_CENTER


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _group_total(commission: Commission, jobs_type: str, rol: str) -> float:
    for g in (commission.comgroups or []):
        if g.Jobs_type == jobs_type and g.Rol == rol:
            return g.Total_detail or 0.0
    return 0.0


def _reimbursement_total(commission: Commission) -> float:
    return sum(r.Value or 0.0 for r in (commission.reimbursements or []))


def _collect_details(
    commissions: List[Commission], jobs_type: str, rol: str
) -> List[tuple]:
    """Return list of (member_name, group, detail) for the given type+role."""
    rows = []
    for c in commissions:
        member_name = c.member.Member_Name if c.member else ""
        for g in (c.comgroups or []):
            if g.Jobs_type == jobs_type and g.Rol == rol:
                for d in (g.comdetails or []):
                    rows.append((member_name, c.Month, c.Year, d))
    return rows


# ─── Main generator ──────────────────────────────────────────────────────────

def generate_commission_excel(
    session: Session,
    member_ids: Optional[List[str]],
    year: Optional[int],
    month: Optional[str],
) -> bytes:
    stmt = (
        select(Commission)
        .options(
            selectinload(Commission.member),
            selectinload(Commission.reimbursements),
            selectinload(Commission.comgroups)
            .selectinload(CommissionGroup.comdetails)
            .selectinload(CommissionDetail.job)
            .selectinload(Job.client),
        )
        .order_by(Commission.Year.desc(), Commission.Month, Commission.ID_Member)
    )

    if member_ids:
        stmt = stmt.where(Commission.ID_Member.in_(member_ids))
    if year is not None:
        stmt = stmt.where(Commission.Year == year)
    if month:
        stmt = stmt.where(Commission.Month.ilike(month))

    commissions: List[Commission] = session.exec(stmt).unique().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Commissions"

    # ── Section 1: Summary ────────────────────────────────────────────────────
    SUMMARY_COLS = 11
    DETAIL_COLS = 9

    row = 1
    _section_header(ws, row, "Commission Summary", SUMMARY_COLS, _FILL_SUMMARY)
    row += 1

    summary_headers = [
        "Member", "Year", "Month",
        "QID Sell ($)", "QID Mgmt ($)", "PAR Mgmt ($)",
        "Total Commission ($)", "Reimbursements ($)", "Total Margin ($)",
        "Status", "Applicable",
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        _hdr(ws, row, col_idx, header, _FILL_SUMMARY)
    row += 1

    for c in commissions:
        member_name = c.member.Member_Name if c.member else ""
        qid_sell = _group_total(c, "QID", "Acc Rep Selling")
        qid_mgmt = _group_total(c, "QID", "Mgmt Member")
        par_mgmt = _group_total(c, "PAR", "Mgmt Member")
        reimb = _reimbursement_total(c)

        values = [
            member_name,
            c.Year,
            c.Month,
            qid_sell,
            qid_mgmt,
            par_mgmt,
            c.Total_commission,
            reimb,
            c.Total_margin,
            c.Status,
            "Yes" if c.Applicable else "No",
        ]
        for col_idx, val in enumerate(values, start=1):
            _body(ws, row, col_idx, val)
        row += 1

    # ── Section 2: QID Sell Detail ────────────────────────────────────────────
    row += 1
    _section_header(ws, row, "QID — Acc Rep Selling", DETAIL_COLS, _FILL_SEC_QID_SELL)
    row += 1

    detail_headers = [
        "Member", "GQM ID", "Client", "Month", "Year",
        "Premium ($)", "Margin (%)", "Commission ($)", "Type",
    ]
    for col_idx, header in enumerate(detail_headers, start=1):
        _hdr(ws, row, col_idx, header, _FILL_QID_SELL)
    row += 1

    for member_name, month_val, year_val, d in _collect_details(commissions, "QID", "Acc Rep Selling"):
        job = d.job
        client_name = job.client.Client_Community if (job and job.client) else ""
        premium = job.Gqm_final_prem_in_money if job else None
        margin = job.Gqm_final_percentage if job else None
        values = [
            member_name,
            job.ID_Jobs if job else "",
            client_name,
            month_val,
            year_val,
            premium,
            margin,
            d.Sell_Mgmt,
            _fmt(d.Type),
        ]
        for col_idx, val in enumerate(values, start=1):
            _body(ws, row, col_idx, val)
        row += 1

    # ── Section 3: QID Mgmt Detail ────────────────────────────────────────────
    row += 1
    _section_header(ws, row, "QID — Mgmt Member", DETAIL_COLS, _FILL_SEC_QID_MGMT)
    row += 1

    for col_idx, header in enumerate(detail_headers, start=1):
        _hdr(ws, row, col_idx, header, _FILL_QID_MGMT)
    row += 1

    for member_name, month_val, year_val, d in _collect_details(commissions, "QID", "Mgmt Member"):
        job = d.job
        client_name = job.client.Client_Community if (job and job.client) else ""
        premium = job.Gqm_final_prem_in_money if job else None
        margin = job.Gqm_final_percentage if job else None
        values = [
            member_name,
            job.ID_Jobs if job else "",
            client_name,
            month_val,
            year_val,
            premium,
            margin,
            d.Sell_Mgmt,
            _fmt(d.Type),
        ]
        for col_idx, val in enumerate(values, start=1):
            _body(ws, row, col_idx, val)
        row += 1

    # ── Section 4: PAR Mgmt Detail ────────────────────────────────────────────
    row += 1
    _section_header(ws, row, "PAR — Mgmt Member", DETAIL_COLS, _FILL_SEC_PAR_MGMT)
    row += 1

    for col_idx, header in enumerate(detail_headers, start=1):
        _hdr(ws, row, col_idx, header, _FILL_PAR_MGMT)
    row += 1

    for member_name, month_val, year_val, d in _collect_details(commissions, "PAR", "Mgmt Member"):
        job = d.job
        client_name = job.client.Client_Community if (job and job.client) else ""
        premium = job.Gqm_final_prem_in_money if job else None
        margin = job.Gqm_final_percentage if job else None
        values = [
            member_name,
            job.ID_Jobs if job else "",
            client_name,
            month_val,
            year_val,
            premium,
            margin,
            d.Sell_Mgmt,
            _fmt(d.Type),
        ]
        for col_idx, val in enumerate(values, start=1):
            _body(ws, row, col_idx, val)
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        1: 20, 2: 14, 3: 20, 4: 12, 5: 8,
        6: 14, 7: 12, 8: 16, 9: 12, 10: 12, 11: 12,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
